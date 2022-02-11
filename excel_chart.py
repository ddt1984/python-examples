import openpyxl
from openpyxl.chart import BarChart, Reference

import collections

filename = "samples/data2.xlsx"
workbook = openpyxl.load_workbook(filename)
sheet = workbook.active
first_row = 2 #1번째 줄은 해더 정보

# 2번째 줄부터 한 줄씩 4번째 컬럼(Job Family) 값을 하나씩 증가시키면서 보관
jobFamilyCount = collections.defaultdict(int)
for row in range(first_row, sheet.max_row + 1):
    jf = sheet.cell(row, 4).value
    jobFamilyCount[jf] += 1

# 차트 저장용 시트를 하나 생성
sheet2 = workbook.create_sheet("Job Family Chart")
sheet2.append(("Job Family", "Count")) # 첫번째 줄에 테이블 해더 추가

# 보관해둔 job family별 숫자를 입력
for jf in jobFamilyCount:
    sheet2.append((jf, jobFamilyCount[jf]))

# 차트 데이터용 범위를 정함
categories = Reference(sheet2, min_col=1, min_row=2, max_row=sheet2.max_row)
values = Reference(sheet2, min_col=2, min_row=1, max_col=2, max_row=sheet2.max_row)

# 바 차트 생성하고 D2 필드에 추가 
chart = BarChart()
chart.title = "Job Family Status"
chart.add_data(values, titles_from_data=True)
chart.set_categories(categories)
sheet2.add_chart(chart, "D2")

workbook.save("samples/data2_chart.xlsx")
