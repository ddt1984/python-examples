# https://openpyxl.readthedocs.io/en/stable/

import datetime
import openpyxl

# 엑셀에서 내용 읽기
filename = "samples/data1.xlsx"
workbook = openpyxl.load_workbook(filename)
sheet = workbook.active

# Name 컬럼 값이 "X-Mas"이면 출력
for row in range(1, sheet.max_row): # range() 첫번째 파라미터를 0으로 하면 첫번째 줄부터 시작
    name = sheet.cell(row+1, 1).value
    if name == "X-Mas": 
        score = sheet.cell(row+1, 2).value
        column3 = sheet.cell(row+1, 3).value
        column4 = sheet.cell(row+1, 4).value
        print(name)
        print(score)
        print(column3)
        print(column4)

# Score 컬럼 총합 구하기
total = 0
for row in range(1, sheet.max_row):
    score = sheet.cell(row+1, 2).value
    total = total + score
print("Total score is", score)

# 특정 필드 값 구하기
field = sheet['D3'].value
print("D3 value is", field)

# 마지막 줄에 입력하기
last_row = sheet.max_row + 1
sheet.cell(last_row, 1, "test")
sheet.cell(last_row, 2, 123)
sheet.cell(last_row, 3, datetime.datetime(2022, 1, 1))
sheet.cell(last_row, 3).number_format = 'dd/mm/yyyy'
sheet.cell(last_row, 4, "테스트")

# 마지막에 줄 입력 다른 방법
sheet.append(("test2", 35, datetime.datetime(2022, 1, 1), "테스트으"))
sheet.cell(sheet.max_row, 3).number_format = 'dd/mm/yyyy'

workbook.save("samples/data1_copied.xlsx") # data1_copied.xlsx 파일을 엑셀로 열고 있으면 저장이 안 되니 닫고 실행
